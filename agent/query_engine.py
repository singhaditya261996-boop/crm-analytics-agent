"""
agent/query_engine.py — Core query pipeline for the CRM Analytics Agent.

Pipeline
--------
User question → Intent Classification → Context Building → Code Generation
→ Safe Execution (w/ retry) → Result Formatting → Interpretation
→ Recommendation → Confidence Scoring → QueryResult

Backward-compatible with SelfImprover which imports QueryResult + _execute_code.

Async-primary API
-----------------
engine = QueryEngine(llm_client, dataframes)
result = await engine.run("Top 10 accounts by revenue?")
result = engine.query("…")   # sync wrapper
"""
from __future__ import annotations

import asyncio
import json
import logging
import re
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml

logger = logging.getLogger(__name__)

# ─── Intent types ─────────────────────────────────────────────────────────────

INTENT_TYPES: dict[str, str] = {
    "aggregation":    "sum, count, average, total across a dimension",
    "comparison":     "compare two or more groups, periods, or segments",
    "ranking":        "top N, bottom N, sorted list, league table",
    "trend":          "over time, month on month, year on year",
    "pivot":          "breakdown of X by Y, matrix view, cross-tab",
    "recommendation": "should we, what should, prioritise, focus on",
    "what_if":        "if we, scenario, hypothetical, what would happen",
    "data_quality":   "missing, null, incomplete, duplicate, error",
    "benchmark":      "good, normal, typical, industry, compare to",
}

# Intent types that benefit from benchmark injection
_BENCHMARK_INTENTS = frozenset({"benchmark", "aggregation", "comparison", "trend", "ranking"})

# ─── Prompt templates ─────────────────────────────────────────────────────────

_CLASSIFY_SYSTEM = """\
You are a CRM query intent classifier.
Classify the user question into exactly one category. Respond with ONLY the
category name — one word, lowercase, no punctuation.

Categories:
{intent_descriptions}

Valid responses: {valid_intents}"""

_CODE_SYSTEM = """\
You are a pandas expert analysing CRM sales data for Equans,
a UK facilities management company.

Question intent: {intent_type}
Available dataframes: {table_names}

Schemas:
{schema_context}

Data quality notes:
{quality_notes}

Pattern from memory (use as reference if relevant):
{pattern_memory}

Write a single Python code block to answer the question.

Rules:
- Only use pandas (pd) and numpy (np) — no other imports
- Store final result in variable called 'result'
- result must be a DataFrame, scalar, or dict
- Handle nulls gracefully — never let NaN propagate silently
- For pivot questions: use pd.pivot_table(), store in 'result'
- For ranking questions: include a 'rank' column, sort descending by default
- For trend questions: ensure date column is parsed and sorted correctly
- For what-if questions: create a dict with 'baseline' and 'scenario' keys
- For benchmark questions: focus on computing the metric (benchmarks injected separately)
- No print statements
- If the question cannot be answered with available data:
  result = 'CANNOT_ANSWER: <specific reason>'"""

_INTERPRET_SYSTEM = """\
You are a senior sales analytics consultant analysing CRM data
for Equans, a UK facilities management company.

The user asked: {question}
Intent type: {intent_type}
The analysis returned: {result_summary}
Industry benchmarks (if relevant): {benchmark_context}

Write a clear, concise business interpretation. Follow this structure:

HEADLINE: One sentence stating the single most important finding.

CONTEXT: 1-2 sentences explaining what this means for the business.
If benchmark data is available, compare the result against it explicitly:
e.g. "This is above/below the FM sector median of X%"

DETAIL: 1-2 sentences on any notable patterns, outliers, or nuances.

Keep total length under 150 words.
Use business language not technical language.
Never mention pandas, dataframes, or code."""

_RECOMMEND_SYSTEM = """\
You are a senior strategy consultant. Based on this CRM analysis result
for Equans, provide a brief actionable recommendation.

Analysis result: {result_summary}
Interpretation: {interpretation_text}
Intent type: {intent_type}

Produce exactly three items in this exact format:

\U0001f3af PRIORITY ACTION: [one specific sentence naming accounts/service lines/deal sizes]

\u26a0\ufe0f RISK FLAG: [one sentence on the most significant risk if nothing changes]

\U0001f4a1 OPPORTUNITY: [one sentence on the clearest growth or improvement opportunity]

Be direct and commercial. Avoid generic advice."""

_FIX_CODE_PROMPT = """\
The previous code failed with the following error. Rewrite it to fix the issue.

Previous code:
```python
{code}
```

Error:
{error}

Write corrected Python code only. Store the result in a variable called 'result'."""


# ─── QueryResult ──────────────────────────────────────────────────────────────

@dataclass
class QueryResult:
    """
    Structured result from the query pipeline.

    Original fields (kept for backward compatibility with SelfImprover):
        question, code, result, error, iterations, chart_hint

    Extended fields added by Module 6 pipeline:
        intent_type, answer_text, recommendation, result_df, pivot_df,
        chart, confidence_score, benchmark_used, final_score,
        provider_used, timestamp, session_id
    """
    # ── Original fields (SelfImprover depends on these) ───────────────────────
    question: str
    code: str              # pandas code that was generated and executed
    result: Any            # raw execution result (DataFrame | scalar | str | None)
    error: str | None = None
    iterations: int = 1
    chart_hint: str | None = None   # set by chart_generator later

    # ── Extended pipeline fields ───────────────────────────────────────────────
    intent_type: str = "aggregation"
    answer_text: str = ""            # plain English business interpretation
    recommendation: dict = field(default_factory=dict)  # {priority_action, risk_flag, opportunity}
    result_df: pd.DataFrame | None = None   # formatted result for display
    pivot_df: pd.DataFrame | None = None    # populated for pivot intent only
    chart: Any = None                       # plotly Figure (set by chart module)
    confidence_score: int = 0              # 0–100
    benchmark_used: bool = False
    final_score: int = 0                   # after self-improvement
    provider_used: str = ""
    timestamp: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    session_id: str = ""

    # ── Convenience aliases ────────────────────────────────────────────────────
    @property
    def code_used(self) -> str:
        return self.code

    @property
    def iterations_taken(self) -> int:
        return self.iterations

    @property
    def confidence_label(self) -> str:
        if self.confidence_score >= 85:
            return "\U0001f7e2 High confidence"
        if self.confidence_score >= 60:
            return "\U0001f7e1 Review recommended"
        return "\U0001f534 Low confidence \u2014 verify manually"


# ─── Module-level executors (backward compat) ─────────────────────────────────

def safe_execute(
    code: str, dataframes: dict[str, pd.DataFrame]
) -> tuple[Any, str | None]:
    """
    Execute LLM-generated pandas code in an isolated namespace.

    Returns (result, error_message). error_message is None on success.
    Provides pd, np, and all dataframes in namespace.
    """
    namespace: dict[str, Any] = {"pd": pd, "np": np, **dataframes}
    try:
        exec(compile(code, "<llm_generated>", "exec"), namespace)  # noqa: S102
        result = namespace.get("result")
        if result is None:
            return "CANNOT_ANSWER: No result variable found in generated code", None
        return result, None
    except Exception as exc:
        return None, f"{type(exc).__name__}: {str(exc)}\n{traceback.format_exc()}"


def _execute_code(code: str, dataframes: dict[str, pd.DataFrame]) -> Any:
    """
    Module-level executor shared with SelfImprover.
    Raises on any error (unlike safe_execute which returns error string).
    """
    namespace: dict[str, Any] = {"pd": pd, "np": np, **dataframes}
    exec(compile(code, "<llm_generated>", "exec"), namespace)  # noqa: S102
    if "result" not in namespace:
        raise NameError("LLM code did not assign a variable named `result`.")
    return namespace["result"]


def _extract_code(llm_output: str) -> str:
    """Extract the first Python code block from LLM output (backward compat name)."""
    m = re.search(r"```(?:python)?\s*(.*?)```", llm_output, re.DOTALL)
    if m:
        return m.group(1).strip()
    return llm_output.strip()


def _sanitise_code(code: str) -> str:
    """Reject obviously dangerous patterns (backward compat)."""
    banned = [
        "import os", "import sys", "subprocess", "open(",
        "__import__", "eval(", "exec(",
    ]
    for pattern in banned:
        if pattern in code:
            raise ValueError(f"Unsafe code pattern detected: {pattern!r}")
    return code


# ─── IntentClassifier ─────────────────────────────────────────────────────────

class IntentClassifier:
    """Classifies user questions into one of the 9 INTENT_TYPES via LLM."""

    _KEYWORD_HINTS: dict[str, list[str]] = {
        "aggregation":    ["total", "sum", "average", "count", "how many", "how much"],
        "comparison":     ["compare", "versus", "vs", "difference", "between"],
        "ranking":        ["top", "bottom", "best", "worst", "highest", "lowest", "ranked"],
        "trend":          ["over time", "trend", "month", "year", "quarterly", "growth"],
        "pivot":          ["breakdown", "cross", "matrix", "pivot"],
        "recommendation": ["should", "recommend", "prioritise", "focus", "what to"],
        "what_if":        ["if we", "scenario", "what would", "hypothetical", "what if"],
        "data_quality":   ["missing", "null", "incomplete", "duplicate", "error"],
        "benchmark":      ["benchmark", "industry", "typical", "normal", "sector"],
    }

    def __init__(self, llm_client: Any) -> None:
        self._llm = llm_client

    async def classify(self, question: str) -> str:
        """Return the intent type string for the question."""
        intent_descriptions = "\n".join(
            f"  {name}: {desc}" for name, desc in INTENT_TYPES.items()
        )
        valid_intents = ", ".join(INTENT_TYPES)
        system = _CLASSIFY_SYSTEM.format(
            intent_descriptions=intent_descriptions,
            valid_intents=valid_intents,
        )
        try:
            raw = await self._llm.complete(system, question)
            intent = raw.strip().lower().split()[0] if raw.strip() else ""
            if intent in INTENT_TYPES:
                return intent
        except Exception as exc:
            logger.warning("Intent classification LLM call failed: %s", exc)
        return self._keyword_fallback(question)

    def _keyword_fallback(self, question: str) -> str:
        q = question.lower()
        for intent, keywords in self._KEYWORD_HINTS.items():
            for kw in keywords:
                if re.search(r"\b" + re.escape(kw) + r"\b", q):
                    return intent
        return "aggregation"


# ─── RecommendationEngine ─────────────────────────────────────────────────────

class RecommendationEngine:
    """Generates a 'So What?' section (priority action, risk flag, opportunity)."""

    def __init__(self, llm_client: Any) -> None:
        self._llm = llm_client

    async def generate(
        self,
        result_summary: str,
        interpretation: str,
        intent_type: str,
    ) -> dict[str, str]:
        system = _RECOMMEND_SYSTEM.format(
            result_summary=result_summary,
            interpretation_text=interpretation,
            intent_type=intent_type,
        )
        try:
            raw = await self._llm.complete(system, "Generate the recommendation.")
            return self._parse(raw)
        except Exception as exc:
            logger.warning("Recommendation LLM call failed: %s", exc)
            return {"priority_action": "", "risk_flag": "", "opportunity": ""}

    @staticmethod
    def _parse(text: str) -> dict[str, str]:
        """Parse structured recommendation text into a dict."""
        out: dict[str, str] = {"priority_action": "", "risk_flag": "", "opportunity": ""}
        patterns = {
            "priority_action": re.compile(
                r"(?:\U0001f3af|🎯)\s*PRIORITY ACTION\s*:?\s*(.+?)(?=(?:\u26a0\ufe0f|⚠️|\U0001f4a1|💡)|$)",
                re.S | re.I,
            ),
            "risk_flag": re.compile(
                r"(?:\u26a0\ufe0f|⚠️)\s*RISK FLAG\s*:?\s*(.+?)(?=(?:\U0001f3af|🎯|\U0001f4a1|💡)|$)",
                re.S | re.I,
            ),
            "opportunity": re.compile(
                r"(?:\U0001f4a1|💡)\s*OPPORTUNITY\s*:?\s*(.+?)(?=(?:\U0001f3af|🎯|\u26a0\ufe0f|⚠️)|$)",
                re.S | re.I,
            ),
        }
        for key, pattern in patterns.items():
            m = pattern.search(text)
            if m:
                out[key] = m.group(1).strip()
        # Plain-text fallback (no emojis)
        if not any(out.values()):
            for key, label in [
                ("priority_action", "PRIORITY ACTION"),
                ("risk_flag", "RISK FLAG"),
                ("opportunity", "OPPORTUNITY"),
            ]:
                m = re.search(
                    rf"{label}\s*:?\s*(.+?)(?=PRIORITY ACTION|RISK FLAG|OPPORTUNITY|$)",
                    text, re.S | re.I,
                )
                if m:
                    out[key] = m.group(1).strip()
        return out


# ─── ConfidenceScorer ─────────────────────────────────────────────────────────

class ConfidenceScorer:
    """Computes a 0–100 confidence score for a query result."""

    BASE = 100
    RETRY_PENALTY = 20     # deducted per retry needed
    NULL_PENALTY_MAX = 30  # max deduction for null-heavy results
    SINGLE_ROW_PENALTY = 10
    ASSUMPTION_PENALTY = 10

    def score(
        self,
        result: Any,
        retries_needed: int,
        has_assumptions: bool = False,
    ) -> int:
        if isinstance(result, str) and result.startswith("CANNOT_ANSWER"):
            return 0
        s = self.BASE - retries_needed * self.RETRY_PENALTY
        if isinstance(result, pd.DataFrame):
            if result.empty:
                s = min(s, 20)   # cap at 20 for empty df; retry penalty still applies
            else:
                null_ratio = result.isnull().mean().mean()
                if not np.isnan(null_ratio):
                    s -= int(null_ratio * self.NULL_PENALTY_MAX)
                if len(result) == 1:
                    s -= self.SINGLE_ROW_PENALTY
        if has_assumptions:
            s -= self.ASSUMPTION_PENALTY
        return max(0, min(100, int(s)))


# ─── PivotHandler ─────────────────────────────────────────────────────────────

class PivotHandler:
    """Formats pivot DataFrames and exports to Excel with formatting."""

    _HEADER_HEX = "1F3864"   # dark blue
    _HEADER_FONT = "FFFFFF"  # white
    _ALT_ROW_HEX = "E8F0FE"  # light blue

    # Currency and percentage column keyword detection
    _CURRENCY_KEYWORDS = frozenset({
        "revenue", "amount", "value", "cost", "price", "deal", "budget",
    })
    _PCT_KEYWORDS = frozenset({
        "pct", "rate", "percent", "ratio", "win_rate",
    })

    def export_to_excel(
        self,
        pivot_df: pd.DataFrame,
        raw_df: pd.DataFrame | None,
        output_dir: Path,
        metadata: dict | None = None,
    ) -> Path:
        """
        Export a pivot table to an xlsx file with:
          - Bold coloured header row
          - Alternating row colours
          - Auto-fitted column widths
          - Optional 'Raw Data' second sheet
          - Optional 'Metadata' third sheet (when *metadata* is provided)
        Returns the output file path.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("Install openpyxl: pip install openpyxl") from exc

        output_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = output_dir / f"pivot_{ts}.xlsx"

        wb = Workbook()
        ws_pivot = wb.active
        ws_pivot.title = "Pivot"
        self._write_sheet(ws_pivot, pivot_df, themed_header=True)

        if raw_df is not None:
            ws_raw = wb.create_sheet("Raw Data")
            self._write_sheet(ws_raw, raw_df, themed_header=True)

        if metadata:
            ws_meta = wb.create_sheet("Metadata")
            meta_rows = [
                ("Query",       metadata.get("question", "")),
                ("Code",        metadata.get("code", "")),
                ("Timestamp",   metadata.get("timestamp", "")),
                ("Data Source", metadata.get("session_id", "")),
                ("Confidence",  metadata.get("confidence_score", "")),
                ("Provider",    metadata.get("provider_used", "")),
            ]
            from openpyxl.styles import Font as _Font
            for r_idx, (key, val) in enumerate(meta_rows, start=1):
                ws_meta.cell(r_idx, 1, value=key).font = _Font(bold=True)
                ws_meta.cell(r_idx, 2, value=str(val))
            ws_meta.column_dimensions["A"].width = 16
            ws_meta.column_dimensions["B"].width = 80

        wb.save(out_path)
        logger.info("Pivot exported to %s", out_path)
        return out_path

    def _write_sheet(
        self, ws: Any, df: pd.DataFrame, themed_header: bool = True
    ) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor=self._HEADER_HEX)
        alt_fill = PatternFill("solid", fgColor=self._ALT_ROW_HEX)
        header_font = Font(color=self._HEADER_FONT, bold=True)

        # Flatten MultiIndex columns if present
        flat = df.copy()
        if isinstance(flat.columns, pd.MultiIndex):
            flat.columns = [
                " | ".join(str(c) for c in col).strip() for col in flat.columns
            ]
        flat = flat.reset_index()
        n_cols = len(flat.columns)

        # Header row
        for c_idx, col_name in enumerate(flat.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col_name))
            if themed_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # Data rows with alternating fill
        for r_idx, row_tuple in enumerate(flat.itertuples(index=False), start=2):
            row_fill = alt_fill if r_idx % 2 == 0 else None
            for c_idx in range(n_cols):
                val = row_tuple[c_idx]
                cell = ws.cell(row=r_idx, column=c_idx + 1, value=val)
                if row_fill:
                    cell.fill = row_fill

        # Number formatting — currency and percentage columns
        for c_idx, col_name in enumerate(flat.columns, start=1):
            col_lower = col_name.lower()
            if any(kw in col_lower for kw in self._CURRENCY_KEYWORDS):
                fmt = '\u00a3#,##0'   # £#,##0
            elif any(kw in col_lower for kw in self._PCT_KEYWORDS):
                fmt = '0.0%'
            else:
                fmt = None
            if fmt and pd.api.types.is_numeric_dtype(flat.iloc[:, c_idx - 1]):
                for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                    for cell in row:
                        cell.number_format = fmt

        # Auto-fit column widths (sample up to 100 rows)
        sample_n = min(len(flat), 100)
        for c_idx in range(n_cols):
            col_name = str(flat.columns[c_idx])
            max_len = len(col_name)
            for r_idx in range(sample_n):
                val_str = str(flat.iloc[r_idx, c_idx])
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = min(
                max_len + 2, 40
            )


# ─── WhatIfHandler ────────────────────────────────────────────────────────────

class WhatIfHandler:
    """Formats what-if scenario results into comparison DataFrames."""

    def format_comparison(self, result: Any) -> pd.DataFrame:
        """
        Convert a what-if result into a Baseline vs Scenario comparison DataFrame.

        Expected input: dict with 'baseline' and 'scenario' keys (each a dict or scalar).
        Falls back gracefully for other types.
        """
        if isinstance(result, pd.DataFrame):
            return result

        if not isinstance(result, dict):
            return pd.DataFrame({"result": [str(result)]})

        baseline = result.get("baseline", {})
        scenario = result.get("scenario", {})

        if isinstance(baseline, dict) and isinstance(scenario, dict):
            keys = sorted(set(baseline) | set(scenario))
            rows = []
            for key in keys:
                b_val = baseline.get(key)
                s_val = scenario.get(key)
                delta: Any = ""
                pct: str = ""
                if isinstance(b_val, (int, float)) and isinstance(s_val, (int, float)):
                    delta = s_val - b_val
                    if b_val != 0:
                        pct = f"{(delta / b_val) * 100:+.1f}%"
                rows.append(
                    {"Metric": key, "Baseline": b_val, "Scenario": s_val,
                     "Change": delta, "% Change": pct}
                )
            return pd.DataFrame(rows)

        # Scalar comparison
        return pd.DataFrame(
            {"Metric": ["Value"], "Baseline": [baseline], "Scenario": [scenario]}
        )


# ─── BenchmarkInjector ────────────────────────────────────────────────────────

class BenchmarkInjector:
    """Loads benchmarks.yaml and matches relevant benchmarks to the question."""

    def __init__(self, benchmarks_path: Path | None = None) -> None:
        self._path = benchmarks_path or (
            Path(__file__).parent.parent / "knowledge" / "benchmarks.yaml"
        )
        self._data: dict = {}
        self._load()

    def _load(self) -> None:
        if self._path.exists():
            try:
                with open(self._path) as f:
                    self._data = yaml.safe_load(f) or {}
            except Exception as exc:
                logger.warning("Could not load benchmarks.yaml: %s", exc)

    @property
    def is_loaded(self) -> bool:
        return bool(self._data)

    def find_relevant(
        self,
        intent_type: str,
        question: str,
        result_df: pd.DataFrame | None = None,
    ) -> str:
        if not self._data:
            return ""
        q_lower = question.lower()
        lines: list[str] = []
        for category, benchmarks in self._data.items():
            if not isinstance(benchmarks, dict):
                continue
            cat_words = category.replace("_", " ").lower().split()
            if any(w in q_lower for w in cat_words):
                for metric, value in benchmarks.items():
                    lines.append(f"  {metric}: {value}")
        return "\n".join(lines) if lines else ""


# ─── QueryEngine ──────────────────────────────────────────────────────────────

class QueryEngine:
    """
    Orchestrates the full query pipeline from plain-English question to QueryResult.

    Usage
    -----
    engine = QueryEngine(llm_client, {"sales": df})
    result = await engine.run("What are our top accounts?")
    result = engine.query("…")   # sync wrapper
    """

    _MAX_RETRIES = 3
    _GRACEFUL_FAILURE = (
        "I wasn't able to compute this \u2014 the data may not contain the fields needed. "
        "Try rephrasing or check the Data Explorer tab."
    )

    def __init__(
        self,
        llm_client: Any,
        dataframes: dict[str, pd.DataFrame],
        schemas: list | None = None,
        quality_report: dict | None = None,
        benchmarks_path: Path | None = None,
        chroma_collection: Any | None = None,
        self_improver: Any | None = None,
        session_id: str | None = None,
        exports_dir: Path | None = None,
        # Legacy compat kwargs
        join_map: dict | None = None,
    ) -> None:
        self._llm = llm_client
        self.dataframes = dict(dataframes)
        self.schemas = schemas or []
        self.quality_report = quality_report or {}
        self.join_map = join_map or {}

        self._classifier = IntentClassifier(llm_client)
        self._recommender = RecommendationEngine(llm_client)
        self._confidence_scorer = ConfidenceScorer()
        self._pivot_handler = PivotHandler()
        self._whatif_handler = WhatIfHandler()
        self._benchmark_injector = BenchmarkInjector(benchmarks_path)
        self._chroma = chroma_collection
        self._self_improver = self_improver

        self.session_id = session_id or str(uuid.uuid4())[:8]
        self.exports_dir = exports_dir or Path("exports")
        self._executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="qe_sync")

        # Legacy compat — used by old-style context_builder import path
        try:
            from agent.context_builder import ContextBuilder as _CB
            self.context_builder = _CB(self.dataframes)
        except Exception:
            self.context_builder = None  # type: ignore[assignment]

    # ── Main async entry point ────────────────────────────────────────────────

    async def run(self, question: str) -> QueryResult:
        """Full async pipeline. Returns a populated QueryResult."""

        # 1. Classify intent
        intent_type = await self._classifier.classify(question)
        logger.info("Intent=%s  Q=%.80s", intent_type, question)

        # 2. Load benchmarks if relevant
        benchmark_context = ""
        benchmark_used = False
        if intent_type in _BENCHMARK_INTENTS:
            benchmark_context = self._benchmark_injector.find_relevant(
                intent_type, question
            )
            benchmark_used = bool(benchmark_context)

        # 3. Build context strings
        schema_ctx = self._build_schema_context()
        quality_notes = self._build_quality_notes()
        pattern_memory = self._query_chroma(question) if self._chroma else "None."
        table_names = ", ".join(self.dataframes) or "none"

        context: dict[str, str] = {
            "intent_type": intent_type,
            "table_names": table_names,
            "schema_context": schema_ctx,
            "quality_notes": quality_notes,
            "pattern_memory": pattern_memory,
        }

        # 4. Generate code + execute (with up to _MAX_RETRIES retries)
        code, raw_result, error_count = await self._generate_and_execute(
            question, context
        )

        # 5. Format result
        result_df, pivot_df = self._format_result(raw_result, intent_type)

        # 6. Build result summary for LLM prompt injection
        result_summary = self._summarise(result_df if result_df is not None else raw_result)

        # 7. Interpret result
        answer_text = await self._interpret(
            question, intent_type, result_summary, benchmark_context
        )

        # 8. Generate recommendation
        recommendation = await self._recommender.generate(
            result_summary, answer_text, intent_type
        )

        # 9. Score confidence
        retries = max(0, error_count)
        confidence_score = self._confidence_scorer.score(
            result_df if result_df is not None else raw_result, retries
        )

        # 10. Self-improvement hook (full impl in Module 13)
        final_score = confidence_score
        if self._self_improver is not None:
            try:
                legacy_qr = QueryResult(
                    question=question, code=code, result=raw_result
                )
                improved = self._self_improver.improve(legacy_qr, self.dataframes)
                if improved and improved.result is not None:
                    final_score = getattr(improved, "confidence_score", confidence_score)
            except Exception as exc:
                logger.warning("Self-improver error (non-fatal): %s", exc)

        return QueryResult(
            question=question,
            code=code,
            result=raw_result,
            error=None,
            iterations=error_count + 1,
            intent_type=intent_type,
            answer_text=answer_text,
            recommendation=recommendation,
            result_df=result_df,
            pivot_df=pivot_df,
            confidence_score=confidence_score,
            benchmark_used=benchmark_used,
            final_score=final_score,
            provider_used=getattr(self._llm, "provider", "unknown"),
            timestamp=datetime.now(timezone.utc),
            session_id=self.session_id,
        )

    def query(self, question: str) -> QueryResult:
        """Sync wrapper for run(). Blocks until the result is ready."""
        return self._run_sync(self.run(question))

    # ── Code generation + execution ───────────────────────────────────────────

    async def _generate_and_execute(
        self, question: str, context: dict[str, str]
    ) -> tuple[str, Any, int]:
        """
        Generate pandas code, execute it safely, retry up to _MAX_RETRIES times.
        Returns (code, result, error_count).
        """
        system = _CODE_SYSTEM.format(**context)
        code = ""
        last_error = ""
        error_count = 0

        for attempt in range(self._MAX_RETRIES + 1):
            if attempt == 0:
                user_prompt = question
            else:
                user_prompt = _FIX_CODE_PROMPT.format(code=code, error=last_error)

            raw_response = await self._llm.complete(system, user_prompt)
            code = _extract_code(raw_response)

            result, error_msg = safe_execute(code, self.dataframes)

            if error_msg is None:
                return code, result, error_count

            last_error = error_msg
            error_count += 1
            logger.warning(
                "Code execution failed (attempt %d/%d): %s",
                attempt + 1, self._MAX_RETRIES + 1, error_msg[:120],
            )

        # All retries exhausted
        logger.error("All %d code generation attempts failed for: %.80s", self._MAX_RETRIES + 1, question)
        return code, self._GRACEFUL_FAILURE, error_count

    # ── Result formatting ──────────────────────────────────────────────────────

    def _format_result(
        self, result: Any, intent_type: str
    ) -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
        """Format raw result into (result_df, pivot_df)."""

        if isinstance(result, str):
            if result.startswith(("CANNOT_ANSWER", "I wasn't able")):
                return pd.DataFrame({"message": [result]}), None
            return pd.DataFrame({"message": [result]}), None

        if intent_type == "pivot" and isinstance(result, pd.DataFrame):
            return result, result

        if intent_type == "what_if":
            formatted = self._whatif_handler.format_comparison(result)
            return formatted, None

        if intent_type == "ranking" and isinstance(result, pd.DataFrame):
            df = result.copy()
            if "rank" not in df.columns:
                df.insert(0, "rank", range(1, len(df) + 1))
            return df, None

        if isinstance(result, pd.DataFrame):
            return result, None

        if isinstance(result, dict):
            try:
                return pd.DataFrame([result]), None
            except Exception:
                return pd.DataFrame({"result": [str(result)]}), None

        if isinstance(result, (int, float)):
            return pd.DataFrame({"result": [result]}), None

        if result is not None:
            return pd.DataFrame({"result": [str(result)]}), None

        return None, None

    # ── Interpretation ────────────────────────────────────────────────────────

    async def _interpret(
        self,
        question: str,
        intent_type: str,
        result_summary: str,
        benchmark_context: str,
    ) -> str:
        if result_summary.startswith(("CANNOT_ANSWER", "I wasn't able")):
            return result_summary
        system = _INTERPRET_SYSTEM.format(
            question=question,
            intent_type=intent_type,
            result_summary=result_summary,
            benchmark_context=benchmark_context or "No benchmark data available.",
        )
        try:
            return await self._llm.complete(system, "Write the business interpretation.")
        except Exception as exc:
            logger.warning("Interpretation LLM call failed: %s", exc)
            return f"Analysis complete. {result_summary}"

    # ── Context building helpers ───────────────────────────────────────────────

    def _build_schema_context(self) -> str:
        if self.schemas:
            lines: list[str] = []
            for schema in self.schemas:
                lines.append(f"Table: {schema.filename} ({schema.row_count} rows)")
                for col in schema.columns:
                    sample = col.sample_values[:3] if col.sample_values else []
                    lines.append(
                        f"  {col.name} ({col.inferred_type}) — sample: {sample}"
                    )
            return "\n".join(lines)
        # Fallback: infer from DataFrames
        lines = []
        for name, df in self.dataframes.items():
            lines.append(f"Table: {name} ({len(df)} rows)")
            for col in df.columns:
                null_pct = df[col].isna().mean() * 100
                sample = df[col].dropna().head(3).tolist()
                lines.append(
                    f"  {col} ({df[col].dtype}) — nulls: {null_pct:.0f}% — sample: {sample}"
                )
        return "\n".join(lines)

    def _build_quality_notes(self) -> str:
        if not self.quality_report:
            return "No quality report available."
        issues = self.quality_report.get("issues", [])
        if not issues:
            return "No data quality issues detected."
        lines = [f"Data quality issues ({len(issues)} total):"]
        for issue in issues[:10]:
            sev = issue.get("severity", "WARNING")
            desc = issue.get("description", "")
            lines.append(f"  [{sev}] {desc}")
        return "\n".join(lines)

    def _query_chroma(self, question: str) -> str:
        try:
            results = self._chroma.query(query_texts=[question], n_results=3)
            docs = results.get("documents", [[]])[0]
            return "\n---\n".join(docs) if docs else "None."
        except Exception:
            return "None."

    # ── Result summarisation ───────────────────────────────────────────────────

    @staticmethod
    def _summarise(result: Any) -> str:
        if result is None:
            return "No result."
        if isinstance(result, str):
            return result
        if isinstance(result, pd.DataFrame):
            if result.empty:
                return "Empty DataFrame."
            preview = result.head(5).to_string(index=False, max_colwidth=50)
            return f"DataFrame ({len(result)} rows, columns: {list(result.columns)}):\n{preview}"
        if isinstance(result, dict):
            return json.dumps(result, default=str)[:500]
        return str(result)[:500]

    # ── Sync helper ────────────────────────────────────────────────────────────

    def _run_sync(self, coro: Any) -> Any:
        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            loop = None
        if loop and loop.is_running():
            future = self._executor.submit(asyncio.run, coro)
            return future.result()
        return asyncio.run(coro)


# ─── Streamlit UI helper ──────────────────────────────────────────────────────

def render_result_ui(result: QueryResult) -> None:
    """Render a QueryResult in Streamlit (sidebar + main area)."""
    try:
        import streamlit as st
    except ImportError as exc:
        raise ImportError("streamlit is required for render_result_ui()") from exc

    # Confidence badge
    st.caption(result.confidence_label)

    # Main interpretation
    st.markdown(result.answer_text)

    # Result dataframe
    if result.result_df is not None and not result.result_df.empty:
        st.dataframe(result.result_df, use_container_width=True)

    # Pivot export
    if result.pivot_df is not None:
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Export as Excel Pivot"):
                handler = PivotHandler()
                out_path = handler.export_to_excel(
                    result.pivot_df,
                    result.result_df,
                    Path("exports"),
                )
                st.success(f"Saved: {out_path.name}")
                with open(out_path, "rb") as fh:
                    st.download_button(
                        "Download xlsx", fh, file_name=out_path.name
                    )

    # Recommendation panel
    rec = result.recommendation
    if any(rec.values()):
        with st.expander("\U0001f4cb Recommendations — So What?", expanded=False):
            st.markdown(
                f"""
<div style="background:#0e7c7b;padding:1rem;border-radius:6px;color:white">

**\U0001f3af Priority Action:** {rec.get('priority_action','')}

**\u26a0\ufe0f Risk Flag:** {rec.get('risk_flag','')}

**\U0001f4a1 Opportunity:** {rec.get('opportunity','')}

</div>""",
                unsafe_allow_html=True,
            )

    # Code expander
    with st.expander("View generated code", expanded=False):
        st.code(result.code, language="python")
