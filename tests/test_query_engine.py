"""
tests/test_query_engine.py — Unit tests for agent/query_engine.py (Module 6).

All LLM calls are mocked. No Ollama/Groq required.
Covers: IntentClassifier, RecommendationEngine, ConfidenceScorer,
        PivotHandler, WhatIfHandler, BenchmarkInjector, safe_execute,
        _execute_code, _extract_code, _sanitise_code, QueryEngine pipeline.
"""
from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any

import pandas as pd
import numpy as np
import pytest

from agent.query_engine import (
    INTENT_TYPES,
    BenchmarkInjector,
    ConfidenceScorer,
    IntentClassifier,
    PivotHandler,
    QueryEngine,
    QueryResult,
    RecommendationEngine,
    WhatIfHandler,
    _execute_code,
    _extract_code,
    _sanitise_code,
    safe_execute,
)


# ─── Mock LLM helpers ─────────────────────────────────────────────────────────

class MockLLMClient:
    """Sequential async mock — each call returns the next response in the list."""

    def __init__(self, responses: list[str] | str = "aggregation") -> None:
        self._responses = responses if isinstance(responses, list) else [responses]
        self._call_count = 0
        self.provider = "mock"

    async def complete(self, system: str, user: str, temperature: Any = None) -> str:
        idx = min(self._call_count, len(self._responses) - 1)
        self._call_count += 1
        return self._responses[idx]

    # Sync compat for self_improver / old tests
    def chat(self, messages: list, temperature: Any = None) -> str:
        return asyncio.run(self.complete("", ""))


def _make_full_mock(
    intent: str = "aggregation",
    code: str = "result = sales['revenue'].sum()",
    interpretation: str = "HEADLINE: Revenue is £1M.\n\nCONTEXT: Good.\n\nDETAIL: None.",
    recommendation: str = (
        "\U0001f3af PRIORITY ACTION: Focus on Hard FM.\n"
        "\u26a0\ufe0f RISK FLAG: Pipeline low.\n"
        "\U0001f4a1 OPPORTUNITY: Upsell to top 5 accounts."
    ),
) -> MockLLMClient:
    """Return a mock whose 4 calls match the full pipeline sequence."""
    code_block = f"```python\n{code}\n```"
    return MockLLMClient([intent, code_block, interpretation, recommendation])


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture()
def sales_df() -> pd.DataFrame:
    return pd.DataFrame({
        "account_id": [1, 2, 3, 4, 5],
        "account_name": ["Acme", "Beta", "Gamma", "Delta", "Epsilon"],
        "revenue": [100_000, 250_000, 75_000, 500_000, 320_000],
        "stage": ["Closed Won", "Prospecting", "Closed Won", "Negotiation", "Closed Won"],
        "service_line": ["Hard FM", "Soft FM", "Hard FM", "Soft FM", "Hard FM"],
    })


@pytest.fixture()
def engine(sales_df, tmp_path) -> QueryEngine:
    llm = _make_full_mock()
    return QueryEngine(
        llm_client=llm,
        dataframes={"sales": sales_df},
        exports_dir=tmp_path / "exports",
    )


@pytest.fixture()
def benchmark_yaml(tmp_path) -> Path:
    content = """win_rate:\n  fm_median: "28% (BIFM 2024)"\nsales_cycle_days:\n  median: "84 days"\n"""
    p = tmp_path / "benchmarks.yaml"
    p.write_text(content)
    return p


# ─── TestQueryResult ──────────────────────────────────────────────────────────

class TestQueryResult:
    def test_backward_compat_fields_accessible(self):
        qr = QueryResult(question="q", code="result = 1", result=1)
        assert qr.question == "q"
        assert qr.code == "result = 1"
        assert qr.result == 1
        assert qr.error is None
        assert qr.iterations == 1

    def test_code_used_alias(self):
        qr = QueryResult(question="q", code="x = 1", result=None)
        assert qr.code_used == "x = 1"

    def test_iterations_taken_alias(self):
        qr = QueryResult(question="q", code="", result=None, iterations=3)
        assert qr.iterations_taken == 3

    def test_confidence_label_high(self):
        qr = QueryResult(question="q", code="", result=None, confidence_score=90)
        assert "High" in qr.confidence_label

    def test_confidence_label_medium(self):
        qr = QueryResult(question="q", code="", result=None, confidence_score=70)
        assert "Review" in qr.confidence_label

    def test_confidence_label_low(self):
        qr = QueryResult(question="q", code="", result=None, confidence_score=40)
        assert "Low" in qr.confidence_label

    def test_default_recommendation_is_empty_dict(self):
        qr = QueryResult(question="q", code="", result=None)
        assert qr.recommendation == {}

    def test_extended_fields_have_defaults(self):
        qr = QueryResult(question="q", code="", result=None)
        assert qr.intent_type == "aggregation"
        assert qr.answer_text == ""
        assert qr.confidence_score == 0
        assert qr.benchmark_used is False


# ─── TestSafeExecute ─────────────────────────────────────────────────────────

class TestSafeExecute:
    def test_success_returns_result_and_no_error(self, sales_df):
        code = "result = sales['revenue'].sum()"
        val, err = safe_execute(code, {"sales": sales_df})
        assert err is None
        assert val == sales_df["revenue"].sum()

    def test_result_is_dataframe(self, sales_df):
        code = "result = sales.head(3)"
        val, err = safe_execute(code, {"sales": sales_df})
        assert err is None
        assert isinstance(val, pd.DataFrame)
        assert len(val) == 3

    def test_error_returns_none_result_and_error_string(self, sales_df):
        code = "result = nonexistent_df['col']"
        val, err = safe_execute(code, {"sales": sales_df})
        assert val is None
        assert err is not None
        assert "NameError" in err

    def test_no_result_variable_returns_cannot_answer(self, sales_df):
        code = "x = 42"
        val, err = safe_execute(code, {"sales": sales_df})
        assert err is None
        assert isinstance(val, str)
        assert "CANNOT_ANSWER" in val

    def test_syntax_error_returned_as_error_string(self, sales_df):
        code = "result = !!!"
        val, err = safe_execute(code, {"sales": sales_df})
        assert val is None
        assert err is not None

    def test_numpy_available_in_namespace(self, sales_df):
        code = "result = np.array([1, 2, 3]).sum()"
        val, err = safe_execute(code, {"sales": sales_df})
        assert err is None
        assert val == 6

    def test_multiple_dataframes_accessible(self, sales_df):
        other = pd.DataFrame({"x": [10, 20]})
        code = "result = len(sales) + len(other)"
        val, err = safe_execute(code, {"sales": sales_df, "other": other})
        assert err is None
        assert val == 7


# ─── TestExecuteCode ─────────────────────────────────────────────────────────

class TestExecuteCode:
    def test_returns_result(self, sales_df):
        code = "result = sales['revenue'].max()"
        assert _execute_code(code, {"sales": sales_df}) == 500_000

    def test_raises_name_error_without_result_variable(self, sales_df):
        with pytest.raises(NameError, match="result"):
            _execute_code("x = 1", {"sales": sales_df})

    def test_raises_on_bad_code(self, sales_df):
        with pytest.raises(Exception):
            _execute_code("result = nonexistent_var", {"sales": sales_df})

    def test_numpy_available(self, sales_df):
        code = "result = np.median(sales['revenue'].values)"
        assert _execute_code(code, {"sales": sales_df}) == pytest.approx(250_000)


# ─── TestExtractCode ─────────────────────────────────────────────────────────

class TestExtractCode:
    def test_extracts_python_fence(self):
        raw = "```python\nresult = df.head()\n```"
        assert _extract_code(raw) == "result = df.head()"

    def test_extracts_plain_fence(self):
        raw = "```\nresult = 42\n```"
        assert _extract_code(raw) == "result = 42"

    def test_returns_raw_when_no_fence(self):
        raw = "result = df.head()"
        assert _extract_code(raw) == "result = df.head()"

    def test_trims_whitespace(self):
        raw = "```python\n  result = 1  \n```"
        assert _extract_code(raw) == "result = 1"

    def test_extracts_first_fence_when_multiple(self):
        raw = "```python\nresult = 1\n```\nsome text\n```python\nresult = 2\n```"
        assert _extract_code(raw) == "result = 1"


# ─── TestSanitiseCode ────────────────────────────────────────────────────────

class TestSanitiseCode:
    def test_clean_code_passes(self):
        code = "result = df.groupby('stage')['revenue'].sum()"
        assert _sanitise_code(code) == code

    @pytest.mark.parametrize("bad", [
        "import os; result = os.listdir('/')",
        "import sys; result = sys.argv",
        "result = subprocess.run(['ls'])",
        "result = open('secret.txt').read()",
        "result = eval('1+1')",
    ])
    def test_rejects_unsafe_patterns(self, bad):
        with pytest.raises(ValueError, match="Unsafe code pattern"):
            _sanitise_code(bad)


# ─── TestIntentClassifier ─────────────────────────────────────────────────────

class TestIntentClassifier:
    def _run_classify(self, question: str, llm_response: str) -> str:
        llm = MockLLMClient(llm_response)
        clf = IntentClassifier(llm)
        return asyncio.run(clf.classify(question))

    @pytest.mark.parametrize("intent", list(INTENT_TYPES))
    def test_returns_valid_intent_when_llm_correct(self, intent):
        result = self._run_classify("some question", intent)
        assert result == intent

    def test_keyword_fallback_for_ranking(self):
        # LLM returns garbage → keyword fallback should detect "top"
        llm = MockLLMClient("not_a_valid_intent")
        clf = IntentClassifier(llm)
        result = asyncio.run(clf.classify("Show me the top 5 accounts by revenue"))
        assert result == "ranking"

    def test_keyword_fallback_for_trend(self):
        llm = MockLLMClient("invalid_response")
        clf = IntentClassifier(llm)
        result = asyncio.run(clf.classify("What is the revenue trend over time?"))
        assert result == "trend"

    def test_keyword_fallback_for_what_if(self):
        llm = MockLLMClient("invalid_response")
        clf = IntentClassifier(llm)
        result = asyncio.run(clf.classify("What if we increased win rate by 10%?"))
        assert result == "what_if"

    def test_fallback_to_aggregation_for_unknown(self):
        llm = MockLLMClient("xyz")
        clf = IntentClassifier(llm)
        result = asyncio.run(clf.classify("something completely unrecognisable xyzzy"))
        assert result == "aggregation"

    def test_llm_response_stripped_and_lowercased(self):
        llm = MockLLMClient("  RANKING  ")
        clf = IntentClassifier(llm)
        result = asyncio.run(clf.classify("Top 10 accounts"))
        assert result == "ranking"

    def test_llm_exception_falls_back_to_keyword(self):
        class FailingLLM:
            provider = "mock"
            async def complete(self, system, user, temperature=None):
                raise ConnectionError("offline")
        clf = IntentClassifier(FailingLLM())
        result = asyncio.run(clf.classify("compare Hard FM vs Soft FM revenue"))
        assert result == "comparison"


# ─── TestRecommendationEngine ─────────────────────────────────────────────────

class TestRecommendationEngine:
    _SAMPLE = (
        "\U0001f3af PRIORITY ACTION: Target top 5 accounts.\n"
        "\u26a0\ufe0f RISK FLAG: Pipeline is thin in Q4.\n"
        "\U0001f4a1 OPPORTUNITY: Upsell Hard FM services."
    )

    def test_generate_returns_dict_with_three_keys(self):
        llm = MockLLMClient(self._SAMPLE)
        engine = RecommendationEngine(llm)
        rec = asyncio.run(engine.generate("summary", "interpretation", "ranking"))
        assert set(rec) == {"priority_action", "risk_flag", "opportunity"}

    def test_priority_action_parsed(self):
        rec = RecommendationEngine._parse(self._SAMPLE)
        assert "Target top 5" in rec["priority_action"]

    def test_risk_flag_parsed(self):
        rec = RecommendationEngine._parse(self._SAMPLE)
        assert "Pipeline is thin" in rec["risk_flag"]

    def test_opportunity_parsed(self):
        rec = RecommendationEngine._parse(self._SAMPLE)
        assert "Upsell Hard FM" in rec["opportunity"]

    def test_all_fields_populated_from_plain_text(self):
        plain = (
            "PRIORITY ACTION: Do this.\n"
            "RISK FLAG: Watch out.\n"
            "OPPORTUNITY: Try that."
        )
        rec = RecommendationEngine._parse(plain)
        assert rec["priority_action"] != ""
        assert rec["risk_flag"] != ""
        assert rec["opportunity"] != ""

    def test_llm_failure_returns_empty_dict(self):
        class FailingLLM:
            provider = "mock"
            async def complete(self, system, user, temperature=None):
                raise RuntimeError("down")
        engine = RecommendationEngine(FailingLLM())
        rec = asyncio.run(engine.generate("s", "i", "ranking"))
        assert isinstance(rec, dict)
        assert "priority_action" in rec


# ─── TestConfidenceScorer ─────────────────────────────────────────────────────

class TestConfidenceScorer:
    def test_perfect_score_no_retries_good_df(self, sales_df):
        scorer = ConfidenceScorer()
        s = scorer.score(sales_df, retries_needed=0)
        assert s == 100

    def test_one_retry_deducts_20(self, sales_df):
        scorer = ConfidenceScorer()
        s = scorer.score(sales_df, retries_needed=1)
        assert s == 80

    def test_three_retries_deducts_60(self, sales_df):
        scorer = ConfidenceScorer()
        s = scorer.score(sales_df, retries_needed=3)
        assert s == 40

    def test_empty_dataframe_returns_20(self):
        scorer = ConfidenceScorer()
        s = scorer.score(pd.DataFrame(), retries_needed=0)
        assert s == 20

    def test_cannot_answer_returns_zero(self):
        scorer = ConfidenceScorer()
        s = scorer.score("CANNOT_ANSWER: no column found", retries_needed=0)
        assert s == 0

    def test_single_row_df_penalised(self, sales_df):
        scorer = ConfidenceScorer()
        s_single = scorer.score(sales_df.head(1), retries_needed=0)
        s_multi = scorer.score(sales_df, retries_needed=0)
        assert s_single < s_multi

    def test_null_heavy_df_penalised(self):
        scorer = ConfidenceScorer()
        null_df = pd.DataFrame({"a": [None, None, None, 1], "b": [None, None, 1, 1]})
        s = scorer.score(null_df, retries_needed=0)
        assert s < 100

    def test_assumption_flag_deducts(self, sales_df):
        scorer = ConfidenceScorer()
        s_no = scorer.score(sales_df, retries_needed=0, has_assumptions=False)
        s_yes = scorer.score(sales_df, retries_needed=0, has_assumptions=True)
        assert s_yes < s_no

    def test_score_clamped_to_zero_minimum(self):
        scorer = ConfidenceScorer()
        s = scorer.score(pd.DataFrame(), retries_needed=10)
        assert s == 0

    def test_score_clamped_to_100_maximum(self, sales_df):
        scorer = ConfidenceScorer()
        s = scorer.score(sales_df, retries_needed=0)
        assert s <= 100


# ─── TestPivotHandler ────────────────────────────────────────────────────────

class TestPivotHandler:
    @pytest.fixture()
    def pivot_df(self) -> pd.DataFrame:
        return pd.DataFrame({
            "service_line": ["Hard FM", "Soft FM"],
            "Q1": [100_000, 80_000],
            "Q2": [120_000, 95_000],
        }).set_index("service_line")

    @pytest.fixture()
    def raw_df(self, sales_df) -> pd.DataFrame:
        return sales_df

    def test_export_creates_xlsx_file(self, pivot_df, raw_df, tmp_path):
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, raw_df, tmp_path)
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_xlsx_has_pivot_sheet(self, pivot_df, tmp_path):
        from openpyxl import load_workbook
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, None, tmp_path)
        wb = load_workbook(out)
        assert "Pivot" in wb.sheetnames

    def test_xlsx_has_raw_data_sheet_when_provided(self, pivot_df, raw_df, tmp_path):
        from openpyxl import load_workbook
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, raw_df, tmp_path)
        wb = load_workbook(out)
        assert "Raw Data" in wb.sheetnames

    def test_xlsx_no_raw_data_sheet_when_none(self, pivot_df, tmp_path):
        from openpyxl import load_workbook
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, None, tmp_path)
        wb = load_workbook(out)
        assert "Raw Data" not in wb.sheetnames

    def test_xlsx_header_row_is_bold(self, pivot_df, tmp_path):
        from openpyxl import load_workbook
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, None, tmp_path)
        wb = load_workbook(out)
        ws = wb["Pivot"]
        assert ws.cell(row=1, column=1).font.bold is True

    def test_multiindex_columns_flattened_in_export(self, tmp_path):
        from openpyxl import load_workbook
        arrays = [["Hard FM", "Hard FM", "Soft FM"], ["Q1", "Q2", "Q1"]]
        mi = pd.MultiIndex.from_arrays(arrays, names=["service", "quarter"])
        df = pd.DataFrame([[1, 2, 3]], columns=mi)
        handler = PivotHandler()
        out = handler.export_to_excel(df, None, tmp_path)
        wb = load_workbook(out)
        ws = wb["Pivot"]
        # All header cells should be plain strings (not tuples)
        for cell in ws[1]:
            if cell.value is not None:
                assert isinstance(cell.value, str)

    def test_output_filename_has_timestamp(self, pivot_df, tmp_path):
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, None, tmp_path)
        assert "pivot_" in out.name

    def test_export_creates_output_dir_if_missing(self, pivot_df, tmp_path):
        new_dir = tmp_path / "deep" / "nested"
        handler = PivotHandler()
        out = handler.export_to_excel(pivot_df, None, new_dir)
        assert out.exists()


# ─── TestWhatIfHandler ───────────────────────────────────────────────────────

class TestWhatIfHandler:
    def test_dict_result_becomes_comparison_df(self):
        handler = WhatIfHandler()
        result = {
            "baseline": {"win_rate": 0.28, "revenue": 2_100_000},
            "scenario": {"win_rate": 0.38, "revenue": 2_850_000},
        }
        df = handler.format_comparison(result)
        assert isinstance(df, pd.DataFrame)
        assert "Baseline" in df.columns
        assert "Scenario" in df.columns

    def test_delta_calculated_correctly(self):
        handler = WhatIfHandler()
        result = {
            "baseline": {"revenue": 1_000_000},
            "scenario": {"revenue": 1_500_000},
        }
        df = handler.format_comparison(result)
        row = df[df["Metric"] == "revenue"].iloc[0]
        assert row["Change"] == pytest.approx(500_000)

    def test_pct_change_calculated(self):
        handler = WhatIfHandler()
        result = {
            "baseline": {"win_rate": 0.28},
            "scenario": {"win_rate": 0.38},
        }
        df = handler.format_comparison(result)
        row = df[df["Metric"] == "win_rate"].iloc[0]
        assert "%" in row["% Change"]

    def test_dataframe_result_returned_as_is(self):
        handler = WhatIfHandler()
        df_in = pd.DataFrame({"A": [1, 2]})
        df_out = handler.format_comparison(df_in)
        pd.testing.assert_frame_equal(df_in, df_out)

    def test_scalar_result_wrapped(self):
        handler = WhatIfHandler()
        df = handler.format_comparison("some_string")
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 1


# ─── TestBenchmarkInjector ───────────────────────────────────────────────────

class TestBenchmarkInjector:
    def test_empty_when_file_does_not_exist(self, tmp_path):
        injector = BenchmarkInjector(tmp_path / "nonexistent.yaml")
        assert not injector.is_loaded

    def test_loads_benchmarks_yaml(self, benchmark_yaml):
        injector = BenchmarkInjector(benchmark_yaml)
        assert injector.is_loaded

    def test_find_relevant_returns_matching_benchmarks(self, benchmark_yaml):
        injector = BenchmarkInjector(benchmark_yaml)
        result = injector.find_relevant("benchmark", "What is the win rate?")
        assert "28%" in result

    def test_find_relevant_empty_when_no_keyword_match(self, benchmark_yaml):
        injector = BenchmarkInjector(benchmark_yaml)
        result = injector.find_relevant("aggregation", "show me accounts with zero deals")
        assert result == ""

    def test_is_loaded_false_when_empty(self, tmp_path):
        empty = tmp_path / "empty.yaml"
        empty.write_text("{}")
        injector = BenchmarkInjector(empty)
        assert not injector.is_loaded


# ─── TestQueryEnginePipeline ─────────────────────────────────────────────────

class TestQueryEnginePipeline:
    def test_run_returns_query_result_instance(self, engine):
        qr = asyncio.run(engine.run("Total revenue?"))
        assert isinstance(qr, QueryResult)

    def test_run_sets_intent_type(self, sales_df, tmp_path):
        llm = _make_full_mock(intent="ranking")
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("Top 5 accounts?"))
        assert qr.intent_type == "ranking"

    def test_run_sets_question_field(self, engine):
        q = "What is our average deal size?"
        qr = asyncio.run(engine.run(q))
        assert qr.question == q

    def test_run_sets_answer_text(self, engine):
        qr = asyncio.run(engine.run("Revenue summary"))
        assert len(qr.answer_text) > 0

    def test_run_sets_recommendation(self, engine):
        qr = asyncio.run(engine.run("Revenue summary"))
        assert isinstance(qr.recommendation, dict)
        assert "priority_action" in qr.recommendation

    def test_run_sets_confidence_score(self, engine):
        qr = asyncio.run(engine.run("Revenue summary"))
        assert 0 <= qr.confidence_score <= 100

    def test_run_sets_code_field(self, engine):
        qr = asyncio.run(engine.run("Revenue summary"))
        assert isinstance(qr.code, str)
        assert len(qr.code) > 0

    def test_run_sets_session_id(self, engine):
        qr = asyncio.run(engine.run("Revenue?"))
        assert len(qr.session_id) > 0

    def test_run_sets_provider_used(self, engine):
        qr = asyncio.run(engine.run("Revenue?"))
        assert qr.provider_used == "mock"

    def test_retry_on_code_failure(self, sales_df, tmp_path):
        """LLM returns bad code twice, then good code on third attempt."""
        llm = MockLLMClient([
            "aggregation",                        # intent
            "result = totally_broken()",          # bad code (attempt 1)
            "result = also_broken.what()",        # bad code (attempt 2, retry 1)
            "```python\nresult = sales['revenue'].sum()\n```",  # good code (attempt 3, retry 2)
            "HEADLINE: OK.\n\nCONTEXT: Fine.\n\nDETAIL: None.",  # interpret
            "\U0001f3af PRIORITY ACTION: X.\n\u26a0\ufe0f RISK FLAG: Y.\n\U0001f4a1 OPPORTUNITY: Z.",  # recommend
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("Total revenue?"))
        assert qr.result == pytest.approx(sales_df["revenue"].sum())
        assert qr.iterations == 3  # 2 retries = 3 iterations (error_count=2, iterations=3)

    def test_graceful_failure_after_max_retries(self, sales_df, tmp_path):
        """All code attempts fail — returns graceful failure message."""
        llm = MockLLMClient([
            "aggregation",               # intent
            "result = broken_always()",  # attempt 1
            "result = broken_always()",  # retry 1
            "result = broken_always()",  # retry 2
            "result = broken_always()",  # retry 3
            "HEADLINE: Failed.",          # interpret
            "\U0001f3af PRIORITY ACTION: N/A.\n\u26a0\ufe0f RISK FLAG: N/A.\n\U0001f4a1 OPPORTUNITY: N/A.",
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("This will always fail"))
        assert "wasn't able" in qr.result or "CANNOT_ANSWER" in str(qr.result)

    def test_pivot_intent_populates_pivot_df(self, sales_df, tmp_path):
        pivot_code = (
            "result = pd.pivot_table(sales, values='revenue', "
            "index='service_line', columns='stage', aggfunc='sum', fill_value=0)"
        )
        llm = MockLLMClient([
            "pivot",
            f"```python\n{pivot_code}\n```",
            "HEADLINE: Pivot done.\n\nCONTEXT: OK.\n\nDETAIL: None.",
            "\U0001f3af PRIORITY ACTION: A.\n\u26a0\ufe0f RISK FLAG: B.\n\U0001f4a1 OPPORTUNITY: C.",
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("Breakdown revenue by service line and stage"))
        assert qr.pivot_df is not None
        assert isinstance(qr.pivot_df, pd.DataFrame)

    def test_ranking_intent_adds_rank_column(self, sales_df, tmp_path):
        rank_code = "result = sales.nlargest(3, 'revenue')[['account_name', 'revenue']]"
        llm = MockLLMClient([
            "ranking",
            f"```python\n{rank_code}\n```",
            "HEADLINE: Top 3.\n\nCONTEXT: OK.\n\nDETAIL: None.",
            "\U0001f3af PRIORITY ACTION: A.\n\u26a0\ufe0f RISK FLAG: B.\n\U0001f4a1 OPPORTUNITY: C.",
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("Top 3 accounts by revenue"))
        assert qr.result_df is not None
        assert "rank" in qr.result_df.columns

    def test_what_if_intent_formats_comparison(self, sales_df, tmp_path):
        whatif_code = (
            'result = {"baseline": {"revenue": 1_000_000}, '
            '"scenario": {"revenue": 1_500_000}}'
        )
        llm = MockLLMClient([
            "what_if",
            f"```python\n{whatif_code}\n```",
            "HEADLINE: Scenario.\n\nCONTEXT: OK.\n\nDETAIL: None.",
            "\U0001f3af PRIORITY ACTION: A.\n\u26a0\ufe0f RISK FLAG: B.\n\U0001f4a1 OPPORTUNITY: C.",
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("If win rate goes up 10%?"))
        assert qr.result_df is not None
        assert "Baseline" in qr.result_df.columns
        assert "Scenario" in qr.result_df.columns

    def test_sync_query_wrapper_works(self, engine):
        qr = engine.query("Revenue summary")
        assert isinstance(qr, QueryResult)

    def test_benchmark_used_flag_set_for_benchmark_intent(self, sales_df, tmp_path, benchmark_yaml):
        llm = MockLLMClient([
            "benchmark",
            "```python\nresult = sales['revenue'].mean()\n```",
            "HEADLINE: Benchmark.\n\nCONTEXT: OK.\n\nDETAIL: None.",
            "\U0001f3af PRIORITY ACTION: A.\n\u26a0\ufe0f RISK FLAG: B.\n\U0001f4a1 OPPORTUNITY: C.",
        ])
        eng = QueryEngine(
            llm, {"sales": sales_df},
            benchmarks_path=benchmark_yaml,
            exports_dir=tmp_path,
        )
        qr = asyncio.run(eng.run("How does our win rate compare to the industry?"))
        assert qr.benchmark_used is True

    def test_no_benchmark_used_flag_for_data_quality_intent(self, sales_df, tmp_path):
        llm = MockLLMClient([
            "data_quality",
            "```python\nresult = sales.isnull().sum()\n```",
            "HEADLINE: Nulls.\n\nCONTEXT: OK.\n\nDETAIL: None.",
            "\U0001f3af PRIORITY ACTION: A.\n\u26a0\ufe0f RISK FLAG: B.\n\U0001f4a1 OPPORTUNITY: C.",
        ])
        eng = QueryEngine(llm, {"sales": sales_df}, exports_dir=tmp_path)
        qr = asyncio.run(eng.run("Are there missing values?"))
        assert qr.benchmark_used is False

    def test_result_df_is_dataframe_on_success(self, engine):
        qr = asyncio.run(engine.run("Revenue summary"))
        assert qr.result_df is None or isinstance(qr.result_df, pd.DataFrame)

    def test_timestamp_is_set(self, engine):
        from datetime import datetime
        qr = asyncio.run(engine.run("Revenue summary"))
        assert isinstance(qr.timestamp, datetime)
